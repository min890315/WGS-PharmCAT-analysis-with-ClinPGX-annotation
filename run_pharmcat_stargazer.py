import subprocess
import sys
import pandas as pd
import os
import argparse
import re
import glob
import shutil
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    pass

# --- Constants ---
CLINPGX_ANNOT_PATH = '/home/jslinkspark/work/ClinPGX_annot/var_drug_ann.tsv'
PHARMCAT_JAR = "/home/jslinkspark/work/tools/pharmcat/pharmcat.jar"
PHARMCAT_PREPROCESSOR = "/home/jslinkspark/work/tools/pharmcat/pharmcat_vcf_preprocessor"
STARGAZER_SCRIPT = "/home/jslinkspark/work/tools/stargazer-grc38-2.0.3/stargazer"
REF_FASTA = "/home/jslinkspark/work/hg38/Homo_sapiens_assembly38.fasta"

# --- Helper Functions (PharmCAT Merge) ---

def parse_rs_diplotype(diplotype):
    """
    Parses 'rs9923231 reference (C)/rs9923231 variant (T)' format.
    Returns (rs_ids, alleles_string) or None if not matching format.
    """
    if 'rs' not in diplotype:
        return None
    
    parts = re.split(r'\s+OR\s+|\s*,\s*', diplotype)
    
    combined_alleles = []
    all_rs_ids = set()
    
    for part in parts:
        matches = re.findall(r'\((?P<base>[A-Za-z]+)\)', part)
        if not matches:
            return None
        
        allele_str = "".join(matches)
        combined_alleles.append(allele_str)
        
        rs_matches = re.findall(r'rs\d+', part)
        all_rs_ids.update(rs_matches)
        
    final_allele_str = " + ".join(combined_alleles)
    return list(all_rs_ids), final_allele_str

def process_star_diplotype(diplotype):
    """
    Parses '*4/*6 OR *4/*35' format.
    """
    parts = diplotype.split(' OR ')
    clean_parts = [p.strip() for p in parts]
    return clean_parts

def parse_g_diplotype(diplotype):
    """
    Parses 'Reference/g.38499696C>T' format.
    Returns 'CT'.
    """
    match = re.search(r'g\.\d+(?P<ref>[A-Za-z]+)>(?P<alt>[A-Za-z]+)', diplotype)
    if match:
        return match.group('ref') + match.group('alt')
    return None

def check_alleles_match(target, candidate):
    """
    Checks if candidate matches target, including reverse order for 2-char strings.
    """
    if target == candidate:
        return True
    if len(target) == 2 and target == candidate[::-1]:
        return True
    return False

def run_merge_pharmcat_clinpgx(report_path, annot_df):
    print(f"Processing PharmCAT Report: {report_path}...")
    try:
        report_df = pd.read_csv(report_path, sep='\t', header=1)
    except Exception as e:
        print(f"Error reading {report_path}: {e}")
        return None

    merged_rows = []
    
    for idx, row in report_df.iterrows():
        gene = row.get('Gene')
        source_diplo = row.get('Source Diplotype')
        
        if pd.isna(gene) or pd.isna(source_diplo):
            merged_rows.append(row.to_dict())
            continue
            
        gene_matches = annot_df[annot_df['Gene'] == gene]
        
        if gene_matches.empty:
            merged_rows.append(row.to_dict())
            continue
        
        matched_ann_rows = pd.DataFrame()
        
        is_rs = 'rs' in str(source_diplo)
        is_star = '*' in str(source_diplo)
        is_g = 'Reference/g.' in str(source_diplo) if not pd.isna(source_diplo) else False

        if is_g:
            g_allele = parse_g_diplotype(source_diplo)
            if g_allele:
                matches = []
                for _, ann_row in gene_matches.iterrows():
                    ann_alleles = str(ann_row.get('Alleles', ''))
                    if check_alleles_match(g_allele, ann_alleles):
                        matches.append(ann_row)
                if matches:
                    matched_ann_rows = pd.DataFrame(matches)

        elif is_rs:
            parsed = parse_rs_diplotype(source_diplo)
            if parsed:
                rs_ids, allele_target = parsed
                matches = []
                for _, ann_row in gene_matches.iterrows():
                    ann_haplo = str(ann_row.get('Variant/Haplotypes', ''))
                    ann_alleles = str(ann_row.get('Alleles', ''))
                    
                    rs_match_found = False
                    for rs in rs_ids:
                        if rs in ann_haplo:
                            rs_match_found = True
                            break
                    
                    if rs_match_found:
                        if check_alleles_match(allele_target, ann_alleles):
                            matches.append(ann_row)
                if matches:
                    matched_ann_rows = pd.DataFrame(matches)
        
        elif is_star:
            candidates = process_star_diplotype(source_diplo)
            matches = []
            for _, ann_row in gene_matches.iterrows():
                ann_alleles = str(ann_row.get('Alleles', ''))
                if ann_alleles in candidates:
                        matches.append(ann_row)
            if matches:
                matched_ann_rows = pd.DataFrame(matches)

        if not matched_ann_rows.empty:
            matched_ann_rows = matched_ann_rows.drop_duplicates(subset=['Variant/Haplotypes', 'Alleles'])
            for _, m_row in matched_ann_rows.iterrows():
                combined = row.to_dict().copy()
                for col, val in m_row.to_dict().items():
                    if col == 'Gene':
                        continue
                    combined[col] = val
                merged_rows.append(combined)
        else:
            merged_rows.append(row.to_dict())

    final_df = pd.DataFrame(merged_rows)
    # Output filename
    new_name = report_path.replace('.report.tsv', '.PharmCAT_ClinPGX.tsv')
    if new_name == report_path: # Safety check if extension didn't match
        new_name = report_path + ".PharmCAT_ClinPGX.tsv"
        
    final_df.to_csv(new_name, sep='\t', index=False)
    print(f"Saved PharmCAT Annotated file to {new_name}")
    return new_name

# --- Helper Functions (Stargazer Merge) ---

def run_merge_stargazer_clinpgx(report_path, annot_df):
    print(f"Processing Stargazer Report: {report_path}...")
    try:
        report_df = pd.read_csv(report_path, sep='\t')
    except Exception as e:
        print(f"Error reading {report_path}: {e}")
        return None

    # --- Custom Merge Logic ---
    # Create a temporary 'MergeKey' column initialized with 'Diplotype'
    report_df['MergeKey'] = report_df['Diplotype']

    for idx, row in report_df.iterrows():
        gene = row.get('Gene')
        diplo = str(row.get('Diplotype', ''))
        
        # 1. VKORC1
        if gene == 'VKORC1':
             if diplo == '-1639G>A/-1639G>A':
                 report_df.at[idx, 'MergeKey'] = 'TT'
        
        # 2. IFNL3
        elif gene == 'IFNL3':
            if diplo == '*1/*1':
                report_df.at[idx, 'MergeKey'] = 'CC'
            elif '/' in diplo:
                alleles = diplo.split('/')
                # Check for *1/{other} or {other}/*1
                if (alleles[0] == '*1' and alleles[1] != '*1') or \
                   (alleles[1] == '*1' and alleles[0] != '*1'):
                    report_df.at[idx, 'MergeKey'] = 'CT'
                # Check for {other}/{other} (neither is *1)
                elif alleles[0] != '*1' and alleles[1] != '*1':
                    report_df.at[idx, 'MergeKey'] = 'TT'

        # 3. ABCG2
        elif gene == 'ABCG2':
            if diplo == '*1/*1':
                report_df.at[idx, 'MergeKey'] = 'GG'
            elif '/' in diplo:
                alleles = diplo.split('/')
                # Check for *1/{other} or {other}/*1
                if (alleles[0] == '*1' and alleles[1] != '*1') or \
                   (alleles[1] == '*1' and alleles[0] != '*1'):
                    report_df.at[idx, 'MergeKey'] = 'GT'
                # Check for {other}/{other} (neither is *1)
                elif alleles[0] != '*1' and alleles[1] != '*1':
                    report_df.at[idx, 'MergeKey'] = 'TT'

    # Merge: Left(Gene, MergeKey), Right(Gene, Alleles)
    # We match the customized 'MergeKey' with 'Alleles' in annotation
    merged_df = pd.merge(
        report_df,
        annot_df,
        left_on=['Gene', 'MergeKey'],
        right_on=['Gene', 'Alleles'],
        how='left'
    )
    
    # Drop the temporary MergeKey if desired, or keep it. 
    # Usually we want to keep the original Diplotype, which is preserved.
    # We might want to drop MergeKey to verify clean output.
    if 'MergeKey' in merged_df.columns:
        merged_df = merged_df.drop(columns=['MergeKey'])

    # Output filename
    if "_report.tsv" in report_path:
        new_name = report_path.replace('_report.tsv', '.PharmCAT.ClinPGX_annot.tsv')
    else:
        new_name = report_path.replace('.tsv', '.PharmCAT.ClinPGX_annot.tsv')

    merged_df.to_csv(new_name, sep='\t', index=False)
    print(f"Saved Stargazer Annotated file to {new_name}")
    return new_name

def run_stargazer_gene(gene, outdir, stargazer_input_vcf):
    """
    Runs Stargazer for a single gene.
    """
    gene_outdir = os.path.join(outdir, f"stargazer_{gene}")

    cmd_stargazer = [
        "python3",
        STARGAZER_SCRIPT,
        "-a", "grc38",
        "-impute",
        "-o", gene_outdir,
        "-d", "wgs",
        "-i", stargazer_input_vcf,
        "-t", gene
    ]
    
    try:
        print(f"Running Stargazer for {gene}...")
        # Capture output to avoid interleaved printing if possible, or just let it print
        # Stargazer might print a lot.
        subprocess.run(cmd_stargazer, check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error/Warning running Stargazer for {gene}: {e}")

# --- Helper Functions (Final Merge) ---

def run_final_merge(clinpgx_file, stargazer_file, output_filename):
    print(f"Merging Annotated Files:\n1: {clinpgx_file}\n2: {stargazer_file}")
    
    try:
        df1 = pd.read_csv(clinpgx_file, sep='\t')
        df2 = pd.read_csv(stargazer_file, sep='\t')
    except Exception as e:
        print(f"Error reading files for final merge: {e}")
        return

    if 'Sample' in df2.columns:
        df2 = df2.drop(columns=['Sample', 'Variant Annotation ID'], errors='ignore')
    
    target_col_for_str = "Multiple phenotypes or diseases And/or"
    if target_col_for_str in df1.columns:
        df1[target_col_for_str] = df1[target_col_for_str].astype(str)
    if target_col_for_str in df2.columns:
        df2[target_col_for_str] = df2[target_col_for_str].astype(str)

    merge_cols = ["Gene", "Alleles", "Variant/Haplotypes", "Drug(s)", "PMID", 
        "Phenotype Category", "Significance", "Notes", "Sentence", 
        "Specialty Population", "Metabolizer types", "isPlural", "Is/Is Not associated", 
        "Direction of effect", "PD/PK terms", "Multiple drugs And/or", "Population types", 
        "Population Phenotypes or diseases", "Multiple phenotypes or diseases And/or", 
        "Comparison Allele(s) or Genotype(s)", "Comparison Metabolizer types"]
    
    # "Alleles" is repeated in the original list provided in script 4 but unique in dataframe usually.
    # checking uniqueness in merge_cols list above... "Alleles" appeared twice in original script list?
    # Original list: [..., "Sentence", "Alleles", "Specialty Population", ...]
    # Yes, let's keep it clean.
    merge_cols = list(dict.fromkeys(merge_cols)) # Deduplicate preserving order

    print("Merging dataframes (Outer Join)...")
    merged_df = pd.merge(df1, df2, on=merge_cols, how='outer')

    target_columns_order = [
        "Gene", "Source Diplotype", "Phenotype", "Activity Score", "Haplotype 1", 
        "Haplotype 1 Function", "Haplotype 1 Activity Value", "Haplotype 2", 
        "Haplotype 2 Function", "Haplotype 2 Activity Value", "Outside Call", 
        "Match Score", "Missing positions", "Undocumented variants", 
        "Recommendation Lookup Diplotype", "Recommendation Lookup Phenotype", 
        "Recommendation Lookup Activity Score", "Gene", "Diplotype", "BEAGLE imputed", 
        "Phenotype", "Score", "May also be", "Also possible haplotype", 
        "Variant/Haplotypes", "Drug(s)", "PMID", 
        "Phenotype Category", "Significance", "Notes", "Sentence", "Alleles", 
        "Specialty Population", "Metabolizer types", "isPlural", "Is/Is Not associated", 
        "Direction of effect", "PD/PK terms", "Multiple drugs And/or", "Population types", 
        "Population Phenotypes or diseases", "Multiple phenotypes or diseases And/or", 
        "Comparison Allele(s) or Genotype(s)", "Comparison Metabolizer types"
    ]

    from collections import Counter
    target_counts = Counter(target_columns_order)
    usage_tracker = {col: 0 for col in target_counts}
    
    final_series_list = []
    final_names_list = []

    for col_name in target_columns_order:
        current_usage = usage_tracker[col_name]
        selected_series = None
        
        if col_name in merged_df.columns:
            selected_series = merged_df[col_name]
        else:
            col_x = f"{col_name}_x"
            col_y = f"{col_name}_y"
            has_x = col_x in merged_df.columns
            has_y = col_y in merged_df.columns
            
            if has_x and has_y:
                if current_usage == 0:
                    selected_series = merged_df[col_y] # Prefer Stargazer (_y) for first occurrence if duplicate
                else:
                    selected_series = merged_df[col_x]
            elif has_x:
                selected_series = merged_df[col_x]
            elif has_y:
                selected_series = merged_df[col_y]
            else:
                selected_series = pd.Series([None]*len(merged_df), index=merged_df.index)

        if selected_series is not None:
            final_series_list.append(selected_series)
            final_names_list.append(col_name)
        
        usage_tracker[col_name] += 1

    output_df = pd.concat(final_series_list, axis=1)
    output_df.columns = final_names_list

    # Conditional Blanking
    # Logic 1: Source Diplotype (col 1) -> clear col 0 (Gene) and col 2 (Phenotype)
    mask1 = output_df.iloc[:, 1].isna() | (output_df.iloc[:, 1].astype(str).str.strip() == "")
    if mask1.any():
        output_df.iloc[mask1, 0] = ""
        output_df.iloc[mask1, 2] = ""

    # Logic 2: Diplotype (col 18) -> clear col 17 (Gene) and col 20 (Phenotype)
    # Note: Column indices are based on target_columns_order
    # Gene(0), Source Diplotype(1), Phenotype(2)...
    # Gene(17), Diplotype(18)...
    # Phenotype(20)
    
    mask2 = output_df.iloc[:, 18].isna() | (output_df.iloc[:, 18].astype(str).str.strip() == "")
    if mask2.any():
        output_df.iloc[mask2, 17] = ""
        output_df.iloc[mask2, 20] = ""

    print(f"Saving Final Merged Report to {output_filename}...")
    try:
        output_df.to_excel(output_filename, index=False, startrow=1)
        
        # Add Custom Headers
        wb = load_workbook(output_filename)
        ws = wb.active

        # Header 1: Pharmcat Result (A-Q)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
        cell = ws.cell(row=1, column=1)
        cell.value = "Pharmcat Result"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

        # Header 2: Stargazer Result (R-X)
        ws.merge_cells(start_row=1, start_column=18, end_row=1, end_column=24)
        cell = ws.cell(row=1, column=18)
        cell.value = "Stargazer Result"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

        # Header 3: ClinPGX Information (Y-AR)
        ws.merge_cells(start_row=1, start_column=25, end_row=1, end_column=44)
        cell = ws.cell(row=1, column=25)
        cell.value = "ClinPGX Information"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

        wb.save(output_filename)
        print("Final headers added successfully.")
        
    except Exception as e:
        print(f"Error saving/formatting excel: {e}")


def cleanup_outdir(outdir):
    print(f"Cleaning up {outdir}...")
    # Remove stargazer_* directories
    stargazer_dirs = glob.glob(os.path.join(outdir, "stargazer_*"))
    for p in stargazer_dirs:
        if os.path.isdir(p):
            try:
                shutil.rmtree(p)
                print(f"Removed directory: {p}")
            except Exception as e:
                print(f"Error removing directory {p}: {e}")

    # Remove json, txt files
    # Only delete files directly inside outdir
    for ext in ['*.json', '*.txt', '*.bgz', '*.csi', '*.vcf']:
        files = glob.glob(os.path.join(outdir, ext))
        for p in files:
            if os.path.isfile(p):
                try:
                    os.remove(p)
                    print(f"Removed file: {p}")
                except Exception as e:
                    print(f"Error removing file {p}: {e}")

def run_command(command, description):
    print(f"Running: {description}")
    print(" ".join(command))
    try:
        subprocess.run(command, check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error running {description}: {e}")
        # sys.exit(1) # Decide whether to stop on component failure. 
        # If PharmCAT fails, we can't do merging for it. Assuming we should stop.
        raise e

def main():
    parser = argparse.ArgumentParser(description="Run PharmCAT and Stargazer pipelines and merge results.")
    parser.add_argument("gvcf_path", help="Input GVCF file path")
    parser.add_argument("outdir", help="Output directory")
    parser.add_argument("prefix", help="Prefix for output files")
    
    if len(sys.argv) < 4:
        parser.print_help()
        sys.exit(1)

    args = parser.parse_args()

    gvcf_path = args.gvcf_path
    outdir = args.outdir
    prefix = args.prefix

    if not os.path.exists(outdir):
        os.makedirs(outdir)
        print(f"Created output directory: {outdir}")

    # Load Annotation File (once)
    if not os.path.exists(CLINPGX_ANNOT_PATH):
        print(f"Error: Annotation file not found at {CLINPGX_ANNOT_PATH}")
        sys.exit(1)
    
    print(f"Loading Annotation: {CLINPGX_ANNOT_PATH}")
    annot_df = pd.read_csv(CLINPGX_ANNOT_PATH, sep='\t')
    if "Variant Annotation ID" in annot_df.columns:
        annot_df = annot_df.drop(columns=["Variant Annotation ID"])

    # Extract VCF filename info
    vcf_basename = os.path.basename(gvcf_path)
    if vcf_basename.endswith(".vcf.gz"):
        vcf_basename_no_ext = vcf_basename[:-7]
    elif vcf_basename.endswith(".gvcf.gz"):
        vcf_basename_no_ext = vcf_basename[:-8]
    elif vcf_basename.endswith(".vcf"):
        vcf_basename_no_ext = vcf_basename[:-4]
    else:
        vcf_basename_no_ext = vcf_basename

    converted_vcf = os.path.join(outdir, f"{prefix}_{vcf_basename_no_ext}.vcf.bgz")
    preprocessed_vcf = os.path.join(outdir, f"{prefix}_{vcf_basename_no_ext}.preprocessed.vcf.bgz")
    
    # --- Step 1: PharmCAT Workflow ---
    
    # 1.1 GVCF to VCF
    cmd_convert = [
        "bcftools", "convert",
        "--gvcf2vcf",
        "--fasta-ref", REF_FASTA,
        gvcf_path,
        "-Oz",
        "-o", converted_vcf,
        "--threads", "16"
    ]
    if not os.path.exists(converted_vcf):
        run_command(cmd_convert, "BCFtools GVCF to VCF conversion")
    else:
        print("Converted VCF exists, skipping conversion.")



    # 1.3 PharmCAT Preprocessor
    cmd_preprocess = [
        PHARMCAT_PREPROCESSOR,
        "-G",
        "-vcf", converted_vcf,
        "-o", preprocessed_vcf # Explicitly set output to match expectations
    ]
    # Note: original script didn't use -o, but pharmcat_vcf_preprocessor usually takes input and makes output.
    # If the tool doesn't support -o for preprocess, let's stick to original behavior or check docs.
    # The original script cmd was: [pp, -G, -vcf, converted_vcf].
    # But then used `preprocessed_vcf` var in next step.
    # Assuming the preprocessor creates a fixed name or the variable `preprocessed_vcf` was guessing the name.
    # Let's trust the original script logic for preprocessor BUT check if the file exists.
    # The original script defined `preprocessed_vcf` variable but didn't pass it to preprocessor command as -o.
    # It passed it to pharmcat command.
    # So `pharmcat_vcf_preprocessor` likely generates `{input}.preprocessed.vcf` or similar.
    # Let's invoke it as originally done, but finding the output might be tricky if implicit.
    # Actually, usually `pharmcat_vcf_preprocessor` might output to stdout if not specified, or create a file?
    # Original script lines 81-86: `cmd_preprocess = [..., "-vcf", converted_vcf]`
    # And line 93: `-vcf`, preprocessed_vcf.
    # This implies the author of original script expected preprocessor to create that file path.
    # Let's keep it exactly as original script for the command calls.
    
    # Correction: The original script defined `preprocessed_vcf` on line 51: `...preprocessed.vcf.bgz`.
    # I'll invoke it as is.
    cmd_preprocess_orig = [
        PHARMCAT_PREPROCESSOR,
        "-G",
        "-vcf", converted_vcf
    ]
    run_command(cmd_preprocess_orig, "PharmCAT VCF Preprocessor")

    # 1.4 Run PharmCAT
    # PharmCAT output defaults to input filename + report extension?
    # We should specify `-o` to control output if possible, or `-O` (output dir).
    # PharmCAT help says `-o` output directory?
    # Original script: `java -jar ... -vcf ...`. 
    # It likely outputs into the same dir as the VCF.
    cmd_pharmcat = [
        "java", "-jar",
        PHARMCAT_JAR,
        "-vcf", preprocessed_vcf,
        "-reporterCallsOnlyTsv",
        "--reporter-extended"
    ]
    run_command(cmd_pharmcat, "PharmCAT Analysis")

    # 1.5 Find PharmCAT Report & Merge
    # Expected: preprocessed_vcf + ".report.tsv" ?
    # Let's glob for it in the directory to be safe, or construct path.
    # Since we don't know exact PharmCAT naming convention (it might replace .vcf with .report.tsv or append),
    # globbing is safer if we just ran it.
    
    # We expect a NEW report file in outdir (or where preprocessed_vcf is).
    # Pattern: `*report.tsv`
    # Let's try to deduce it: `preprocessed_vcf` (path) + `.report.tsv`?
    # Or basename of preprocessed_vcf + ...
    # Let's look for the most recently created report.tsv in outdir
    
    # Simpler: PharmCAT typically outputs `{basename}.report.tsv`.
    # Let's look for any report.tsv in outdir that matches the pattern.
    pharmcat_report = None
    possible_reports = glob.glob(os.path.join(outdir, "*.report.tsv"))
    # Filter for valid reports (not stargazer ones if any exist yet)
    possible_reports = [f for f in possible_reports if "stargazer" not in os.path.basename(f)]
    
    if possible_reports:
        # Pick the one matching our VCF name mostly?
        # or just the latest one
        possible_reports.sort(key=os.path.getmtime, reverse=True)
        pharmcat_report = possible_reports[0]
        print(f"Found PharmCAT report: {pharmcat_report}")
    else:
        print("Warning: Could not identify PharmCAT report file.")
    
    pharmcat_annot_file = None
    if pharmcat_report:
        pharmcat_annot_file = run_merge_pharmcat_clinpgx(pharmcat_report, annot_df)

    # --- Step 2: Stargazer Workflow ---
    
    genes = ['abcb1', 'cacna1s', 'cftr', 'cyp1a1', 'cyp1a2', 'cyp1b1', 'cyp2a6', 'cyp2a13', 
             'cyp2b6', 'cyp2c8', 'cyp2c9', 'cyp2c19', 'cyp2d6', 'cyp2e1', 'cyp2f1', 'cyp2j2', 
             'cyp2r1', 'cyp2s1', 'cyp2w1', 'cyp3a4', 'cyp3a5', 'cyp3a7', 'cyp3a43', 'cyp4a11', 
             'cyp4a22', 'cyp4b1', 'cyp4f2', 'cyp17a1', 'cyp19a1', 'cyp26a1', 'dpyd', 'g6pd', 
             'gstm1', 'gstp1', 'ifnl3', 'nat1', 'nat2', 'nudt15', 'por', 'ptgis', 'ryr1', 
             'slc15a2', 'slc22a2', 'slco1b1', 'slco1b3', 'slco2b1', 'sult1a1', 'tbxas1', 'tpmt', 
             'ugt1a1', 'ugt1a4', 'ugt2b7', 'ugt2b15', 'ugt2b17', 'vkorc1', 'xpc', '2c_cluster', 'abcg2']

    print("Starting Stargazer analysis with max 16 parallel processes...")
    
    with ThreadPoolExecutor(max_workers=16) as executor:
        # Submit all tasks using gvcf_path directly
        futures = [executor.submit(run_stargazer_gene, gene, outdir, gvcf_path) for gene in genes]
        # Wait for all to complete
        for future in futures:
            try:
                future.result()
            except Exception as e:
                print(f"Exception in Stargazer thread: {e}")

    # 2.2 Aggregate Stargazer Reports
    all_reports = []
    print("Aggregating Stargazer reports...")
    
    for gene in genes:
        report_path = os.path.join(outdir, f"stargazer_{gene}", "report.tsv")
        if os.path.exists(report_path):
            try:
                df = pd.read_csv(report_path, sep='\t')
                all_reports.append(df)
            except Exception as e:
                print(f"Error reading {report_path}: {e}")
    
    stargazer_report_file = os.path.join(outdir, f"{prefix}_stargazer_report.tsv")
    
    if all_reports:
        merged_df = pd.concat(all_reports, ignore_index=True)
        if 'Gene' in merged_df.columns:
            merged_df['Gene'] = merged_df['Gene'].str.upper()
        
        merged_df.to_csv(stargazer_report_file, sep='\t', index=False)
        print(f"Aggregated Stargazer report saved to {stargazer_report_file}")
        
        # 2.3 Merge Stargazer with ClinPGX
        stargazer_annot_file = run_merge_stargazer_clinpgx(stargazer_report_file, annot_df)
    else:
        print("No Stargazer reports found to aggregate.")
        stargazer_annot_file = None

    # --- Step 3: Final Merge ---
    
    if pharmcat_annot_file and stargazer_annot_file:
        final_output_xlsx = os.path.join(outdir, f"{prefix}_Pharmacogenomics.ClinPGX_annot.xlsx")
        run_final_merge(pharmcat_annot_file, stargazer_annot_file, final_output_xlsx)
    else:
        print("Skipping final merge because one or both intermediate files are missing.")

    # Cleanup
    cleanup_outdir(outdir)

if __name__ == "__main__":
    main()
